import numpy as np
import pandas as pd
import openpyxl

# def read_table(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     # 选择第一个工作表
#     sheet = workbook.active

#     # 初始化空列表，用于存储每列的数据和表头
#     column_data = [[] for _ in range(sheet.max_column)]
#     headers = []

#     # 读取表头
#     for col in range(1, sheet.max_column + 1):
#         headers.append(sheet.cell(row=1, column=col).value)

#     # 遍历每一行，将每列数据添加到对应的列表中
#     for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第2行开始，跳过表头行
#         for col_idx, cell_value in enumerate(row):
#             column_data[col_idx].append(cell_value)

#     # 获取第一列和第二列数据，分别是照片名和物候期
#     first_column = column_data[0]
#     second_column = column_data[1]

#     # 根据照片名分割成不同的拍摄角度
#     part_data = {}
#     for row_idx in range(len(first_column)):
#         cell_value = first_column[row_idx]
#         parts = cell_value.split('_')
#         part_index = parts[2] # 拍摄角度在第二个索引

#         if part_index not in part_data:
#             part_data[part_index] = []

#         part_data[part_index].append([column_data[col_idx][row_idx] for col_idx in range(len(column_data))])

#     # 确定阳面
#     keys_list = list(part_data.keys())
#     part_list = []
#     for i in keys_list:
#         part_list.append(part_data[i])
#     compare_list = [[sublist[1] for sublist in sublist_list] for sublist_list in part_list]

#     # 找到最多照片的角度
#     max_length = max(len(row) for row in compare_list)
    
#     # 根据最多照片数全部补零
#     compare_longest_list = [row + [0] * (max_length - len(row)) for row in compare_list]
    
#     # 转换为矩阵做筛选
#     compare_matrix = np.array(compare_longest_list)
    
#     rows, cols = compare_matrix.shape

#     # 初始化一个数组来存放每行最大的次数
#     wins = np.zeros(rows, dtype=int)

#     # 遍历每列进行比较
#     for col in range(cols):
#         # 找到每列中的最大值索引
#         max_indices = np.argmax(compare_matrix[:, col])
#         # 增加胜出次数（仅当最大值索引与当前行索引相同时）
#         wins[max_indices] += 1

#     max_key = keys_list[np.argmax(wins)]
#     filtered_temp_data = part_data[max_key]

#     # 遍历阳面的数据，进行筛选，不能有物候期的反复
#     filtered_data = []
#     current_max = None

#     for i in range(len(filtered_temp_data)):
#         # 对于第一行，直接加入筛选后的列表中，并设置当前最大值
#         if i == 0:
#             filtered_data.append(filtered_temp_data[i])
#             current_max = filtered_temp_data[i][1]
#         else:
#             # 检查第二列元素是否比当前最大值大
#             if filtered_temp_data[i][1] >= current_max:
#                 filtered_data.append(filtered_temp_data[i])
#                 current_max = filtered_temp_data[i][1]

#     return headers, filtered_data


def read_table(file_path):
    workbook = openpyxl.load_workbook(file_path)
    # 选择第一个工作表
    sheet = workbook.active
    # 初始化空列表，用于存储每行的数据和表头
    row_data = []
    headers = []

    for col in range(1, sheet.max_column + 1):
        headers.append(sheet.cell(row=1, column=col).value)

    # 遍历每一行，将每行数据作为一个列表存储
    for row in sheet.iter_rows(min_row=2, values_only=True):  
        # 从第2行开始，跳过表头行
        row_data.append(list(row))

    return headers, row_data

if __name__ == "__main__":
    # 测试示例
    file_path = r'data\J48.xlsx'
    headers, rows = read_table(file_path)

    # 打印每行数据
    for i in rows:
        print(i)
